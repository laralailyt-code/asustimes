"""
Backfill all commodities' weekly/daily prices from 2024-01-01 to today.

Sources used (in priority):
- Yahoo Finance (yfinance): Gold, Silver, WTI, Brent, Copper (HG=F→USD/t), Aluminum (ALI=F),
  and 8 FX pairs (TWD/CNY/JPY/EUR/BRL/KRW/IDR/INR).
- bot.com.tw BCD: ABS (130041), Corrugated paper (190060), Long-fiber pulp (190020 — only
  through 2025-10-31 to avoid known post-Nov-2025 corruption).
- cnyes JSONP: Cobalt (lcocs), Palladium (PA). API window is ~360d so will only fill
  the last ~12 months, not all the way back to 2024-01-01.
- For items with no free public history (Tin/Nickel/Zinc/Lithium via TE, PC, 黃磷, 鎢):
  leave the existing earliest values in place. The CSV's carry-forward pass will fill
  gaps with '*' tagged carry values (read-only / does not fabricate real data).

Merge policy: source wins for dates it covers; older existing-CSV dates are preserved.
"""

from __future__ import annotations
import csv
import json
import os
import re
import sys
from datetime import datetime, timedelta
from typing import Iterable

CSV_PATH = os.path.join(os.path.dirname(__file__), "2026 Raw material trend history.csv")
START_DATE = "2024-01-01"
END_DATE   = datetime.now().strftime("%Y-%m-%d")

# ---- Item name mapping (must match exact CSV row labels) -------------------
COPPER_MULT = 2204.62  # HG=F is USD/lb → multiply by 2204.62 to get USD/tonne

YF_SYMS: dict[str, tuple[str, float]] = {
    "GC=F":  ("金 (gold) US$/盎司",            1.0),
    "SI=F":  ("銀 (silver) US$/盎司",          1.0),
    "CL=F":  ("石油 西德州 ( US$/桶)",          1.0),
    "BZ=F":  ("石油 北海布蘭特 (US$/桶)",       1.0),
    "HG=F":  ("銅 (copper) US$/tonne",         COPPER_MULT),
    "ALI=F": ("鋁 (aluminum) US$/tonne",       1.0),
    "PA=F":  ("鈀 (palladium) US$/盎司",       1.0),  # NYMEX Pd futures; fills 2024/1/1-2024/12/15 gap where cnyes API ends
    "TWD=X": ("美元 / 台幣",            1.0),
    "CNY=X": ("美元 / 人民幣",          1.0),
    "JPY=X": ("美元 / 日圓",            1.0),
    "EUR=X": ("美元 / 歐元",            1.0),
    "BRL=X": ("美元 / 巴西里爾(巴西幣)", 1.0),
    "KRW=X": ("美元 / 韓圜",            1.0),
    "IDR=X": ("美元 / 印尼盾",          1.0),
    "INR=X": ("美元 / 印度幣",          1.0),
}

BOT_BCD = {
    "130041": ("ABS聚合物(注塑) 中國到岸價 US$/tonne", 1.0,  None),
    "190060": ("瓦楞芯紙 CNY$/tonne",                  1.0,  None),
    # 190020 long-fiber pulp DROPPED entirely: BCD returns ~700 USD/T values across
    # the whole 2024-2026 window, not just post-Nov 2025. Excel + live MoneyDJ are
    # the trustworthy sources. Re-introducing 190020 interleaves bad ~700 USD/T
    # readings with good ~1300-1500 USD/T weekly snapshots.
}

CNYES_FUTURES = {
    "lcocs": ("鈷 (cobalt) US$/tonne",       "https://www.cnyes.com/futures/Javachart/lcocs.html", 20000, 200000),
    "PA":    ("鈀 (palladium) US$/盎司",     "https://www.cnyes.com/futures/html5chart/PA.html",      500, 5000),
}

# Local Excel with weekly historical data (yearly sheets back to 2019)
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "2026 Raw material trend history.xlsx")
EXCEL_SHEETS = ["2024 1~12", "2025 1~12", "2026 1~12"]

def _norm_name(s: str) -> str:
    """Normalize commodity row labels: strip non-breaking spaces and collapse whitespace."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ").replace("　", " ")).strip()

# Map normalized Excel name → exact CSV row label.
# CSV labels include extra spaces in places (e.g. "鎳 (nickel)  US$/tonne") that
# we must preserve verbatim so we hit the existing row instead of creating dupes.
EXCEL_TO_CSV_NAME = {
    "銅 (copper) US$/tonne":           "銅 (copper) US$/tonne",
    "錫 (tin) US$/tonne":              "錫 (tin) US$/tonne",
    "鋁 (aluminum) US$/tonne":         "鋁 (aluminum) US$/tonne",
    "鎳 (nickel) US$/tonne":           "鎳 (nickel)  US$/tonne",   # CSV has 2 spaces
    "鋅 (zinc) US$/tonne":             "鋅 (zinc)  US$/tonne",     # CSV has 2 spaces
    "金 (gold) US$/盎司":              "金 (gold) US$/盎司",
    "銀 (silver) US$/盎司":            "銀 (silver) US$/盎司",
    "鈷 (cobalt) US$/tonne":           "鈷 (cobalt) US$/tonne",
    "鋰 (Lithium) CNY$/tonne":         "鋰 (Lithium) CNY$/tonne",
    "黃磷 CNY$/tonne":                  "黃磷 CNY$/tonne",
    "PC塑料 (SABIC) CNY$/tonne":       "PC塑料 (SABIC) CNY$/tonne",
    "ABS聚合物(注塑) 中國到岸價 US$/tonne": "ABS聚合物(注塑) 中國到岸價 US$/tonne",
    "石油 西德州 ( US$/桶)":            "石油 西德州 ( US$/桶)",
    "石油 北海布蘭特 (US$/桶)":         "石油 北海布蘭特 (US$/桶)",
    "美元 / 台幣":                      "美元 / 台幣",
    "美元 / 人民幣":                    "美元 / 人民幣",
    "美元 / 日圓":                      "美元 / 日圓",
    "美元 / 歐元":                      "美元 / 歐元",
    "美元 / 巴西里爾(巴西幣)":          "美元 / 巴西里爾(巴西幣)",
    "美元 / 韓圜":                      "美元 / 韓圜",
    "美元 / 印尼盾":                    "美元 / 印尼盾",
    "美元 / 印度幣":                    "美元 / 印度幣",
    "NOREXECO 長纖紙漿 USD/T":          "NOREXECO 長纖紙漿  USD/T",  # CSV has 2 spaces
}

def fetch_excel_history() -> dict[str, list[tuple[str, float]]]:
    """Read yearly sheets from local Excel and return {csv_name: [(YYYY-MM-DD, val), ...]}."""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    out: dict[str, list[tuple[str, float]]] = {}
    for sn in EXCEL_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        # Build date list from header (datetime objects)
        dates: list[str | None] = [None]
        for h in header[1:]:
            if isinstance(h, datetime):
                dates.append(h.strftime("%Y-%m-%d"))
            else:
                dates.append(None)
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            raw_name = _norm_name(r[0])
            csv_name = EXCEL_TO_CSV_NAME.get(raw_name)
            if not csv_name:
                continue
            for i, val in enumerate(r[1:], start=1):
                if val is None or i >= len(dates) or dates[i] is None:
                    continue
                # Skip corrupted values like "[object Object]"
                if isinstance(val, str):
                    s = val.strip()
                    if not s or s in ("[object Object]", "N/A", "-", "0"):
                        continue
                    try:
                        fv = float(s.replace(",", ""))
                    except Exception:
                        continue
                else:
                    try:
                        fv = float(val)
                    except Exception:
                        continue
                out.setdefault(csv_name, []).append((dates[i], fv))
    # Dedupe per item, keeping last occurrence
    for k in list(out.keys()):
        d = {}
        for date, v in out[k]:
            d[date] = v
        out[k] = sorted(d.items())
    return out


# ---- Wide-format CSV I/O ---------------------------------------------------
def read_csv() -> tuple[list[str], dict[str, dict[str, str]]]:
    """Return (header_dates_as_YYYY_M_D, {item_name: {date: cell_value}})."""
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    items: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        if not row or not row[0].strip():
            continue
        name = row[0].strip()
        cells: dict[str, str] = {}
        for i, v in enumerate(row[1:], start=1):
            if i < len(header):
                cells[header[i]] = (v or "").strip()
        items[name] = cells
    return header, items


def write_csv(header: list[str], items: dict[str, dict[str, str]]) -> None:
    # Sort header dates chronologically (skip first cell = "項目")
    def _key(s: str):
        try:
            y, m, d = s.split("/")
            return (int(y), int(m), int(d))
        except Exception:
            return (9999, 99, 99)

    dates_sorted = sorted(header[1:], key=_key)
    new_header = ["項目"] + dates_sorted
    out_rows = [new_header]
    for name in sorted(items.keys(), key=lambda x: x):
        cells = items[name]
        row = [name]
        for d in dates_sorted:
            row.append(cells.get(d, ""))
        out_rows.append(row)
    with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerows(out_rows)


def iso_to_csv(date_iso: str) -> str:
    """YYYY-MM-DD → YYYY/M/D (compact form used by the CSV header)."""
    y, m, d = date_iso.split("-")
    return f"{int(y)}/{int(m)}/{int(d)}"


# ---- Source helpers --------------------------------------------------------
def fetch_yf(sym: str, mult: float) -> list[tuple[str, float]]:
    import yfinance as yf
    df = yf.download(sym, start=START_DATE, end=(datetime.strptime(END_DATE, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d"),
                     interval="1d", auto_adjust=True, progress=False, threads=False)
    if df is None or df.empty:
        return []
    # yfinance new column structure: MultiIndex (Price, Ticker)
    if "Close" in df.columns:
        ser = df["Close"]
    else:
        ser = df.iloc[:, 0]
    # If MultiIndex, take first column
    if hasattr(ser, "columns"):
        ser = ser.iloc[:, 0]
    out = []
    for ts, v in ser.dropna().items():
        date = ts.strftime("%Y-%m-%d")
        out.append((date, round(float(v) * mult, 4)))
    return out


def fetch_bot_bcd(code: str) -> list[tuple[str, float]]:
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    url = f"https://fund.bot.com.tw/Z/ZH/ZHG/CZHG.djbcd?A={code}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
        "Referer": "https://fund.bot.com.tw/",
    }
    r = requests.get(url, headers=headers, timeout=20, verify=False)
    data = r.text.strip()
    if not data or len(data) < 20:
        return []
    m = re.search(r"(\d{4}/\d{2}/\d{2})\s+(\d)", data)
    if not m:
        return []
    dates_str = data[:m.start(2)].strip().rstrip(" ")
    vals_str  = data[m.start(2):]
    dates = [d.strip() for d in dates_str.split(",") if re.match(r"\d{4}/\d{2}/\d{2}$", d.strip())]
    vals: list[float] = []
    for v in vals_str.split(","):
        v = v.strip()
        try:
            vals.append(float(v))
        except ValueError:
            break
    return [(d.replace("/", "-"), round(v, 2)) for d, v in zip(dates, vals)]


def fetch_cnyes(code: str, referer: str, vmin: float, vmax: float) -> list[tuple[str, float]]:
    import requests
    url = f"https://www.cnyes.com/futures/highChart/ChartSource.aspx?type=futures&source=javachart&code={code}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": referer,
    }
    r = requests.get(url, headers=headers, timeout=20)
    if r.status_code != 200:
        return []
    text = r.text.strip()
    m = re.match(r"^\((.*)\)\s*;?\s*$", text, re.DOTALL)
    if not m:
        return []
    try:
        payload = json.loads(m.group(1))
    except Exception:
        return []
    # Actual format: list of [timestamp_ms, value] pairs (UTC midnight ts).
    out: list[tuple[str, float]] = []
    if isinstance(payload, list):
        for item in payload:
            if not (isinstance(item, list) and len(item) >= 2):
                continue
            try:
                ts_ms = int(item[0])
                fv = float(item[1])
            except Exception:
                continue
            if not (vmin <= fv <= vmax):
                continue
            iso = datetime.utcfromtimestamp(ts_ms / 1000).strftime("%Y-%m-%d")
            out.append((iso, round(fv, 2)))
    out = sorted({d: v for d, v in out}.items())
    return out


# ---- Main merge ------------------------------------------------------------
def main():
    header, items = read_csv()
    existing_dates_set = set(header[1:])
    new_dates_set: set[str] = set()
    summary: dict[str, dict] = {}

    def _ingest(name: str, points: Iterable[tuple[str, float]], source: str) -> tuple[int, int]:
        if name not in items:
            items[name] = {}
        cells = items[name]
        added = updated = 0
        for iso_date, val in points:
            if iso_date < START_DATE or iso_date > END_DATE:
                continue
            csv_d = iso_to_csv(iso_date)
            new_dates_set.add(csv_d)
            old = cells.get(csv_d, "").rstrip("*").strip()
            new_str = f"{val:.4f}".rstrip("0").rstrip(".") if isinstance(val, float) else str(val)
            # Keep integer-looking values clean
            if old == "":
                cells[csv_d] = new_str
                added += 1
            else:
                # Source wins on overlap
                try:
                    if abs(float(old) - float(new_str)) > 1e-6:
                        cells[csv_d] = new_str
                        updated += 1
                except Exception:
                    cells[csv_d] = new_str
                    updated += 1
        summary[name] = {"source": source, "added": added, "updated": updated, "total_pts": added + updated}
        return added, updated

    print(f"=== Backfill 2024-01-01 → {END_DATE} ===")
    print(f"CSV path: {CSV_PATH}")
    print(f"Existing header date count: {len(existing_dates_set)}")

    # 0. Local Excel (most authoritative for items without free APIs)
    print("\n[0/4] Local Excel weekly history (2024-2026)...")
    try:
        ex_data = fetch_excel_history()
        for csv_name, pts in ex_data.items():
            a, u = _ingest(csv_name, pts, "Excel")
            print(f"  {csv_name:50s} pts={len(pts):4d}  added={a:4d}  updated={u:4d}")
    except Exception as e:
        print(f"  EXCEL ERROR: {type(e).__name__}: {e}")

    # 1. Yahoo Finance
    print("\n[1/3] Yahoo Finance multi-year backfill...")
    for sym, (name, mult) in YF_SYMS.items():
        try:
            pts = fetch_yf(sym, mult)
            a, u = _ingest(name, pts, f"yfinance {sym}")
            print(f"  {sym:6s} → {name:40s} pts={len(pts):4d}  added={a:4d}  updated={u:4d}")
        except Exception as e:
            print(f"  {sym:6s} ERROR: {type(e).__name__}: {e}")

    # 2. bot.com.tw BCD
    print("\n[2/3] bot.com.tw BCD full history...")
    for code, (name, mult, cap_date) in BOT_BCD.items():
        try:
            pts = fetch_bot_bcd(code)
            if cap_date:
                pts = [(d, v) for d, v in pts if d <= cap_date]
            pts = [(d, v * mult) for d, v in pts]
            a, u = _ingest(name, pts, f"bot.com.tw BCD {code}")
            print(f"  {code} → {name:40s} pts={len(pts):4d}  added={a:4d}  updated={u:4d}")
        except Exception as e:
            print(f"  {code} ERROR: {type(e).__name__}: {e}")

    # 3. cnyes JSONP (~360d window, won't reach 2024-01-01)
    print("\n[3/4] cnyes Cobalt + Palladium (360d window)...")
    for code, (name, referer, vmin, vmax) in CNYES_FUTURES.items():
        try:
            pts = fetch_cnyes(code, referer, vmin, vmax)
            a, u = _ingest(name, pts, f"cnyes {code}")
            print(f"  {code:6s} → {name:40s} pts={len(pts):4d}  added={a:4d}  updated={u:4d}")
        except Exception as e:
            print(f"  {code:6s} ERROR: {type(e).__name__}: {e}")

    # 4. Mirror 美元/X → 匯率/X (legacy duplicate rows)
    # The 匯率 / 人民幣, 匯率 / 台幣, 匯率 / 日幣 rows hold the same data as the
    # 美元 / X rows but were started later. Populate them from the now-backfilled
    # 美元 / X cells so the dashboard shows consistent FX history.
    print("\n[4/4] Mirror 美元/X → 匯率/X for CNY, TWD, JPY...")
    fx_mirror = [
        ("美元 / 人民幣", "匯率 / 人民幣"),
        ("美元 / 台幣",   "匯率 / 台幣"),
        ("美元 / 日圓",   "匯率 / 日幣"),
    ]
    for src, dst in fx_mirror:
        if src not in items:
            print(f"  skip {dst}: source {src} missing")
            continue
        if dst not in items:
            items[dst] = {}
        src_cells = items[src]
        dst_cells = items[dst]
        added = updated = 0
        for csv_d, val in src_cells.items():
            if not val.strip() or val.strip() == "0":
                continue
            try:
                # only consider 2024+ dates per backfill scope
                y, m, d = csv_d.split("/")
                if (int(y), int(m), int(d)) < (2024, 1, 1):
                    continue
            except Exception:
                continue
            old = dst_cells.get(csv_d, "").rstrip("*").strip()
            if old == "":
                dst_cells[csv_d] = val
                added += 1
            elif old != val:
                # keep existing - 匯率 row may have manual overrides; do not clobber
                pass
        print(f"  {src} → {dst}  added={added}")

    # 5. Final purge of corrupted residue (after all ingest passes)
    print("\n[5] Purge corrupt cells across all sources...")
    pulp_name = "NOREXECO 長纖紙漿  USD/T"
    if pulp_name in items:
        cleared = 0
        for csv_d, v in list(items[pulp_name].items()):
            s = v.strip().rstrip("*")
            if not s:
                continue
            try:
                fv = float(s.replace(",", ""))
            except Exception:
                continue
            # bot.com.tw 190020 corrupted range = ~705-735 USD/T (also leaks into Excel
            # 2026 sheet). Real long-fiber pulp trades in 1000-1700 USD/T range.
            if fv < 800:
                items[pulp_name][csv_d] = ""
                cleared += 1
        print(f"  {pulp_name}: cleared {cleared} cells with value < 800")
    for nm in ("PC塑料 (SABIC) CNY$/tonne", "PC/ABS塑料 (SABIC) CNY$/tonne"):
        if nm in items:
            n = sum(1 for v in items[nm].values() if v.strip() == "[object Object]")
            for csv_d, v in list(items[nm].items()):
                if v.strip() == "[object Object]":
                    items[nm][csv_d] = ""
            if n:
                print(f"  {nm}: cleared {n} '[object Object]' cells")

    # Merge any new dates into header
    all_dates = set(header[1:]) | new_dates_set
    def _key(s: str):
        try:
            y, m, d = s.split("/")
            return (int(y), int(m), int(d))
        except Exception:
            return (9999, 99, 99)
    new_header = ["項目"] + sorted(all_dates, key=_key)

    # Persist
    print("\n[write] Saving CSV...")
    write_csv(new_header, items)
    print(f"  New header date count: {len(new_header)-1}")

    # Coverage report: for each item, count cells filled from 2024/1/1 onwards
    print("\n=== Coverage from 2024-01-01 onwards ===")
    def is_2024_plus(csv_date: str) -> bool:
        try:
            y, m, d = csv_date.split("/")
            return (int(y), int(m), int(d)) >= (2024, 1, 1)
        except Exception:
            return False
    relevant_dates = [d for d in new_header[1:] if is_2024_plus(d)]
    print(f"  Dates in window: {len(relevant_dates)}")
    for name in sorted(items.keys()):
        filled = sum(1 for d in relevant_dates if items[name].get(d, "").strip() not in ("", "0"))
        flag = "OK" if filled >= len(relevant_dates) * 0.5 else "LOW"
        print(f"  [{flag}] {name:50s} filled={filled}/{len(relevant_dates)}")

    print("\nDone.")


if __name__ == "__main__":
    main()
