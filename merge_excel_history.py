"""Merge real historical prices from Excel into CSV.

Targets only commodities that had problems (1-2 real points + carry-back filling
the whole month with the same value):
- 鈷、錫、鎳、鋅、鋰、黃磷、PC塑料

Excel: 2026 Raw material trend history.xlsx, sheet "2026 1~12"
Format: row 1 = item | dates (datetime objects), row 2+ = item name | values
"""
import csv
from pathlib import Path
from datetime import datetime
import openpyxl

CSV_PATH  = Path(__file__).parent / "2026 Raw material trend history.csv"
XLSX_PATH = Path(__file__).parent / "2026 Raw material trend history.xlsx"

# Excel item name → CSV item name (exact match required)
TARGET_ITEMS = [
    "鈷 (cobalt) US$/tonne",
    "錫 (tin) US$/tonne",
    "鎳 (nickel)  US$/tonne",   # 注意 nickel 後雙空格與 CSV 一致
    "鋅 (zinc)  US$/tonne",     # 同上
    "鋰 (Lithium) CNY$/tonne",
    "黃磷 CNY$/tonne",
    "PC塑料 (SABIC) CNY$/tonne",
]


def date_to_csv_key(d: datetime) -> str:
    return f"{d.year}/{d.month}/{d.day}"


def main():
    # 1. Read Excel
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if "2026 1~12" not in wb.sheetnames:
        print("Sheet '2026 1~12' not found")
        return
    ws = wb["2026 1~12"]

    # Header: row 1, col 2+ are dates
    header_cells = list(ws[1])
    excel_dates = []  # list of (col_index_1based, csv_date_key)
    for i, cell in enumerate(header_cells[1:], start=2):
        v = cell.value
        if isinstance(v, datetime):
            excel_dates.append((i, date_to_csv_key(v)))

    print(f"Excel: {len(excel_dates)} date columns")
    print(f"Range: {excel_dates[0][1]} ~ {excel_dates[-1][1]}")

    # Build map item → {csv_date_key: value} from Excel
    # Normalise: Excel sometimes uses \xa0 (non-breaking space) where CSV uses
    # normal space. Replace before matching/storing.
    def _norm(s: str) -> str:
        return s.replace("\xa0", " ") if isinstance(s, str) else s

    target_set = set(TARGET_ITEMS)
    excel_data = {}  # item -> dict[date_key, value]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        item = _norm(row[0]) if row[0] else None
        if not item or item not in target_set:
            continue
        d = {}
        for col_idx, csv_key in excel_dates:
            v = row[col_idx - 1]  # row is 0-indexed
            if v is None or v == "":
                continue
            try:
                # Strip any commas if string
                if isinstance(v, str):
                    v = float(v.replace(",", "").rstrip("*"))
                else:
                    v = float(v)
                d[csv_key] = v
            except (ValueError, TypeError):
                pass
        excel_data[item] = d
        print(f"  [{item}] Excel has {len(d)} real points")
    wb.close()

    # 2. Read existing CSV
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    item_to_row = {row[0]: i for i, row in enumerate(rows[1:], start=1)}

    # 3. For each target item, write Excel values into CSV
    # Strategy:
    # - If Excel has a date that CSV header doesn't have → add new column
    # - For each Excel date, write the real value (overwrites carry-back '*')
    # - Cells not in Excel and currently '*' tagged → clear them, will be
    #   re-filled by carry-forward later
    csv_dates_set = set(header[1:])
    new_dates = set()
    for item, dates_map in excel_data.items():
        for d in dates_map:
            if d not in csv_dates_set:
                new_dates.add(d)

    if new_dates:
        # Insert new dates in chronological order
        all_dates = sorted(csv_dates_set | new_dates,
                           key=lambda s: datetime.strptime(s, "%Y/%m/%d"))
        new_header = ["項目"] + all_dates
        # Rebuild rows aligned to new header
        rebuilt = [new_header]
        for r in rows[1:]:
            old_data = dict(zip(header[1:], r[1:]))
            new_r = [r[0]] + [old_data.get(d, "") for d in all_dates]
            rebuilt.append(new_r)
        rows = rebuilt
        header = new_header
        item_to_row = {row[0]: i for i, row in enumerate(rows[1:], start=1)}
        print(f"Added {len(new_dates)} new date columns: {sorted(new_dates)}")

    # Now write Excel values into CSV
    written = 0
    cleared_cf = 0
    for item, dates_map in excel_data.items():
        if item not in item_to_row:
            print(f"  [WARN] {item} not in CSV, skipping")
            continue
        ridx = item_to_row[item]
        row = rows[ridx]
        # Pad row
        while len(row) < len(header):
            row.append("")

        # Step A: Clear all carry-forward '*' cells (will be re-applied later)
        for cidx in range(1, len(header)):
            cell = (row[cidx] or "").strip()
            if cell.endswith("*"):
                row[cidx] = ""
                cleared_cf += 1

        # Step B: Write Excel real values
        for d_key, val in dates_map.items():
            cidx = header.index(d_key)
            if val == int(val):
                row[cidx] = str(int(val))
            else:
                row[cidx] = str(round(val, 4))
            written += 1

    print(f"Wrote {written} real values | Cleared {cleared_cf} stale carry-forward cells")

    # 4. Save back
    with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)
    print(f"[DONE] CSV updated: {CSV_PATH}")
    print()
    print("Now re-run apply_carry_forward.py to fill remaining gaps.")


if __name__ == "__main__":
    main()
